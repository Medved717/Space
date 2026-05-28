import pandas as pd
import os
from datetime import datetime
from openpyxl import load_workbook
from openpyxl.styles import Alignment, Font, PatternFill, Border, Side
from openpyxl.utils import get_column_letter

EXCEL_FILE = "fuel_records.xlsx"


def load_data():
    """Загрузка данных из Excel файла"""
    if os.path.exists(EXCEL_FILE):
        try:
            df = pd.read_excel(EXCEL_FILE, dtype={'id': int})
            # Преобразуем в список словарей
            records = df.to_dict('records')
            # Очищаем NaN значения
            for record in records:
                for key, value in record.items():
                    if pd.isna(value):
                        record[key] = ''
            return records
        except Exception as e:
            print(f"Ошибка при загрузке данных: {e}")
            return []
    return []


def save_data(records):
    """Сохранение данных в Excel файл с форматированием"""
    if not records:
        # Создаем пустой DataFrame с нужными колонками
        df = pd.DataFrame(columns=['id', 'fio', 'vehicle', 'datetime',
                                   'fuel_type', 'fuel_1c', 'fuel_nav', 'difference'])
    else:
        df = pd.DataFrame(records)

    # Сохраняем в Excel
    with pd.ExcelWriter(EXCEL_FILE, engine='openpyxl') as writer:
        df.to_excel(writer, sheet_name='Заправки', index=False)

        # Получаем рабочий лист для форматирования
        worksheet = writer.sheets['Заправки']

        # Форматирование заголовков
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=11)
        header_alignment = Alignment(horizontal="center", vertical="center")

        for col in range(1, len(df.columns) + 1):
            cell = worksheet.cell(row=1, column=col)
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment

        # Настройка ширины колонок
        column_widths = {
            'A': 5,  # id
            'B': 25,  # fio
            'C': 20,  # vehicle
            'D': 20,  # datetime
            'E': 12,  # fuel_type
            'F': 12,  # fuel_1c
            'G': 12,  # fuel_nav
            'H': 12  # difference
        }

        for col, width in column_widths.items():
            worksheet.column_dimensions[col].width = width

        # Форматирование чисел и выравнивания
        for row in range(2, len(df) + 2):
            # Выравнивание по центру для всех ячеек
            for col in range(1, len(df.columns) + 1):
                cell = worksheet.cell(row=row, column=col)
                cell.alignment = Alignment(horizontal="center", vertical="center")

            # Специальное форматирование для разницы
            diff_cell = worksheet.cell(row=row, column=8)  # колонка H - difference
            if diff_cell.value is not None:
                try:
                    diff_value = float(diff_cell.value)
                    if diff_value > 0:
                        diff_cell.font = Font(color="FF0000", bold=True)  # Красный для положительной разницы
                    elif diff_value < 0:
                        diff_cell.font = Font(color="008000", bold=True)  # Зеленый для отрицательной
                except:
                    pass

        # Добавление границ
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'),
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )

        for row in worksheet.iter_rows(min_row=1, max_row=len(df) + 1, min_col=1, max_col=len(df.columns)):
            for cell in row:
                cell.border = thin_border


def add_record():
    """Добавление новой записи"""
    print("\n" + "=" * 60)
    print(" ДОБАВЛЕНИЕ НОВОЙ ЗАПРАВКИ")
    print("=" * 60)

    fio = input("ФИО водителя: ").strip()
    vehicle = input("Модель и номер ТС: ").strip()

    # Ввод даты и времени
    while True:
        datetime_str = input("Дата и время (формат: ДД.ММ.ГГГГ ЧЧ:ММ, например: 22.05.2026 14:30): ").strip()
        try:
            # Проверка формата даты
            datetime.strptime(datetime_str, "%d.%m.%Y %H:%M")
            break
        except ValueError:
            print(" Ошибка! Используйте формат: ДД.ММ.ГГГГ ЧЧ:ММ")

    fuel_type = input("Вид заправки (АИ-92, АИ-95, ДТ и т.д.): ").strip()

    # Ввод чисел с проверкой
    while True:
        try:
            fuel_1c = float(input("Количество топлива по 1С (литры): ").replace(',', '.'))
            if fuel_1c <= 0:
                print(" Количество должно быть больше 0!")
                continue
            break
        except ValueError:
            print(" Ошибка! Введите число (например: 45.5)")

    while True:
        try:
            fuel_nav = float(input("Количество топлива по навигации (литры): ").replace(',', '.'))
            if fuel_nav <= 0:
                print(" Количество должно быть больше 0!")
                continue
            break
        except ValueError:
            print(" Ошибка! Введите число (например: 43.2)")

    difference = round(fuel_nav - fuel_1c, 2)

    # Получаем следующий ID
    records = load_data()
    next_id = max([r['id'] for r in records], default=0) + 1

    record = {
        "id": next_id,
        "fio": fio,
        "vehicle": vehicle,
        "datetime": datetime_str,
        "fuel_type": fuel_type,
        "fuel_1c": round(fuel_1c, 2),
        "fuel_nav": round(fuel_nav, 2),
        "difference": difference
    }

    return record


def format_table(records):
    """Форматирование данных в виде таблицы для консоли"""
    if not records:
        return "Нет данных"

    # Заголовки
    headers = ["N", "ФИО", "ТС", "Дата/Время", "Вид", "1С", "Навигация", "Разница"]
    col_widths = [4, 20, 15, 19, 8, 8, 10, 9]

    # Разделитель
    separator = "+" + "+".join(["-" * (w + 2) for w in col_widths]) + "+"

    # Заголовок
    result = [separator]
    header_line = "|"
    for i, header in enumerate(headers):
        header_line += f" {header:<{col_widths[i]}} |"
    result.append(header_line)
    result.append(separator)

    # Данные
    for record in records:
        row = "|"
        row += f" {record['id']:<{col_widths[0]}} |"
        row += f" {record['fio']:<{col_widths[1]}} |"
        row += f" {record['vehicle']:<{col_widths[2]}} |"
        row += f" {record['datetime']:<{col_widths[3]}} |"
        row += f" {record['fuel_type']:<{col_widths[4]}} |"
        row += f" {record['fuel_1c']:>{col_widths[5]}.2f} |"
        row += f" {record['fuel_nav']:>{col_widths[6]}.2f} |"

        # Разница с цветовой индикацией
        diff = record['difference']
        if diff > 0:
            diff_str = f"+{diff:.2f}"
        elif diff < 0:
            diff_str = f"{diff:.2f}"
        else:
            diff_str = f"{diff:.2f}"
        row += f" {diff_str:>{col_widths[7]}} |"

        result.append(row)
        result.append(separator)

    return "\n".join(result)


def show_statistics(records):
    """Показать статистику"""
    if not records:
        print("\n Нет данных для статистики")
        return

    total_1c = sum(r['fuel_1c'] for r in records)
    total_nav = sum(r['fuel_nav'] for r in records)
    total_diff = total_nav - total_1c

    print("\n" + "=" * 60)
    print(" СТАТИСТИКА")
    print("=" * 60)
    print(f"Всего заправок: {len(records)}")
    print(f"Всего по 1С: {total_1c:.2f} л")
    print(f"Всего по навигации: {total_nav:.2f} л")
    print(f"Общая разница (навигация - 1С): {total_diff:+.2f} л")
    print("-" * 60)

    if total_diff > 0:
        print(f" НАВИГАЦИЯ ПОКАЗЫВАЕТ БОЛЬШЕ на {total_diff:.2f} литров")
    elif total_diff < 0:
        print(f" 1С ПОКАЗЫВАЕТ БОЛЬШЕ на {abs(total_diff):.2f} литров")
    else:
        print(" Данные совпадают идеально!")

    # Дополнительная статистика
    if len(records) > 0:
        print("\n Дополнительная статистика:")
        print(f"Среднее значение по 1С: {total_1c / len(records):.2f} л")
        print(f"Среднее значение по навигации: {total_nav / len(records):.2f} л")
        print(f"Средняя разница: {total_diff / len(records):+.2f} л")


def delete_record(records):
    """Удаление записи"""
    if not records:
        print("\n Нет данных для удаления")
        return records

    print("\n" + format_table(records))
    try:
        record_id = int(input("\nВведите номер записи для удаления (0 - отмена): "))
        if record_id == 0:
            return records

        for i, r in enumerate(records):
            if r['id'] == record_id:
                records.pop(i)
                # Перенумеровываем ID
                for j, r in enumerate(records, 1):
                    r['id'] = j
                save_data(records)
                print(f" Запись #{record_id} удалена")
                return records
        print(f" Запись #{record_id} не найдена")
    except ValueError:
        print(" Введите число")

    return records


def edit_record(records):
    """Редактирование записи"""
    if not records:
        print("\n Нет данных для редактирования")
        return records

    print("\n" + format_table(records))
    try:
        record_id = int(input("\nВведите номер записи для редактирования (0 - отмена): "))
        if record_id == 0:
            return records

        for i, r in enumerate(records):
            if r['id'] == record_id:
                print(f"\nРедактирование записи #{record_id}")
                print(f"Текущее ФИО: {r['fio']}")
                new_fio = input("Новое ФИО (Enter - оставить): ").strip()
                if new_fio:
                    r['fio'] = new_fio

                print(f"Текущее ТС: {r['vehicle']}")
                new_vehicle = input("Новое ТС (Enter - оставить): ").strip()
                if new_vehicle:
                    r['vehicle'] = new_vehicle

                print(f"Текущая дата/время: {r['datetime']}")
                new_datetime = input("Новая дата/время (Enter - оставить): ").strip()
                if new_datetime:
                    # Проверка формата даты
                    try:
                        datetime.strptime(new_datetime, "%d.%m.%Y %H:%M")
                        r['datetime'] = new_datetime
                    except ValueError:
                        print(" Ошибка формата даты! Оставлено прежнее значение")

                print(f"Текущий вид топлива: {r['fuel_type']}")
                new_fuel_type = input("Новый вид топлива (Enter - оставить): ").strip()
                if new_fuel_type:
                    r['fuel_type'] = new_fuel_type

                # Пересчёт разницы при изменении топлива
                try:
                    new_fuel_1c = input(f"Текущее по 1С: {r['fuel_1c']}\nНовое (Enter - оставить): ").strip()
                    if new_fuel_1c:
                        r['fuel_1c'] = round(float(new_fuel_1c.replace(',', '.')), 2)

                    new_fuel_nav = input(f"Текущее по навигации: {r['fuel_nav']}\nНовое (Enter - оставить): ").strip()
                    if new_fuel_nav:
                        r['fuel_nav'] = round(float(new_fuel_nav.replace(',', '.')), 2)

                    r['difference'] = round(r['fuel_nav'] - r['fuel_1c'], 2)
                except ValueError:
                    print(" Ошибка ввода чисел!")

                save_data(records)
                print(f" Запись #{record_id} обновлена")
                return records
        print(f" Запись #{record_id} не найдена")
    except ValueError:
        print(" Введите число")

    return records


def search_records(records):
    """Поиск записей"""
    if not records:
        print("\n Нет данных для поиска")
        return

    print("\n" + "=" * 60)
    print(" ПОИСК ЗАПИСЕЙ")
    print("=" * 60)
    print("1. Поиск по ФИО")
    print("2. Поиск по ТС")
    print("3. Поиск по дате")
    print("4. Поиск по виду топлива")

    choice = input("Выберите тип поиска (1-4): ").strip()

    if choice == '1':
        keyword = input("Введите ФИО (часть): ").strip().lower()
        results = [r for r in records if keyword in r['fio'].lower()]
    elif choice == '2':
        keyword = input("Введите модель или номер ТС: ").strip().lower()
        results = [r for r in records if keyword in r['vehicle'].lower()]
    elif choice == '3':
        keyword = input("Введите дату (например: 22.05.2026): ").strip()
        results = [r for r in records if keyword in r['datetime']]
    elif choice == '4':
        keyword = input("Введите вид топлива (например: АИ-95): ").strip().upper()
        results = [r for r in records if keyword in r['fuel_type'].upper()]
    else:
        print(" Неверный выбор")
        return

    if results:
        print(f"\n Найдено записей: {len(results)}")
        print(format_table(results))
    else:
        print("\n Ничего не найдено")


def export_to_csv(records):
    """Экспорт в CSV"""
    if not records:
        print("\n Нет данных для экспорта")
        return

    import csv
    csv_file = f"fuel_records_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    with open(csv_file, 'w', encoding='utf-8-sig', newline='') as f:
        writer = csv.DictWriter(f, fieldnames=['id', 'fio', 'vehicle', 'datetime',
                                               'fuel_type', 'fuel_1c', 'fuel_nav', 'difference'])
        writer.writeheader()
        for r in records:
            writer.writerow(r)
    print(f"\n Данные экспортированы в {csv_file}")
    print(f" Файл сохранен: {os.path.abspath(csv_file)}")


def open_excel_file():
    """Открыть Excel файл (если возможно)"""
    if os.path.exists(EXCEL_FILE):
        try:
            if os.name == 'nt':  # Windows
                os.startfile(EXCEL_FILE)
            elif os.name == 'posix':  # macOS или Linux
                os.system(f'open "{EXCEL_FILE}"' if sys.platform == 'darwin' else f'xdg-open "{EXCEL_FILE}"')
            print(f"\n Открываю файл {EXCEL_FILE}")
        except Exception as e:
            print(f"\n Не удалось открыть файл автоматически: {e}")
            print(f" Файл находится здесь: {os.path.abspath(EXCEL_FILE)}")
    else:
        print("\n Файл Excel еще не создан")


def main():
    records = load_data()

    while True:
        print("\n" + "=" * 60)
        print(" УЧЁТ ЗАПРАВОК ТРАНСПОРТА (Excel)")
        print("=" * 60)
        print("1.  Добавить заправку")
        print("2.  Показать все записи")
        print("3.  Статистика")
        print("4.  Редактировать запись")
        print("5.  Удалить запись")
        print("6.  Поиск записей")
        print("7.  Экспорт в CSV")
        print("8.  Открыть Excel файл")
        print("9.  Выход")
        print("-" * 60)

        choice = input("Выберите действие (1-9): ").strip()

        if choice == '1':
            new_record = add_record()
            records.append(new_record)
            save_data(records)
            print(f"\n Запись добавлена! ID: {new_record['id']}")
            print(f" Разница: {new_record['difference']:+.2f} л")
            print(f" Данные сохранены в {EXCEL_FILE}")

        elif choice == '2':
            if records:
                print("\n" + format_table(records))
                print(f"\n Всего записей: {len(records)}")
            else:
                print("\n Нет записей")

        elif choice == '3':
            show_statistics(records)

        elif choice == '4':
            records = edit_record(records)

        elif choice == '5':
            records = delete_record(records)

        elif choice == '6':
            search_records(records)

        elif choice == '7':
            export_to_csv(records)

        elif choice == '8':
            open_excel_file()

        elif choice == '9':
            print("\n До свидания!")
            print(f" Ваши данные сохранены в файле: {os.path.abspath(EXCEL_FILE)}")
            break

        else:
            print("\n Неверный выбор! Введите 1-9")


if __name__ == "__main__":
    # Проверка наличия необходимых библиотек
    try:
        import pandas
        import openpyxl
    except ImportError as e:
        print("Ошибка: Необходимо установить библиотеки!")
        print("Выполните команду:")
        print("pip install pandas openpyxl")
        exit(1)

    main()